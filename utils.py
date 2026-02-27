import pandas as pd
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
import os
import re
from typing import Optional, Tuple, List, Any, Dict
import time
from urllib.parse import quote
from io import BytesIO

try:
    import streamlit as st
except Exception:
    st = None

try:
    from sqlalchemy import create_engine, text
except Exception:
    create_engine = None
    text = None

EXCEL_FILE = "project.xlsx"

SHEET_DOCUMENT = "Document Master Data"
SHEET_USERS = "Users"
SHEET_FUNGSI = "Master Fungsi"
SHEET_AUDIT = "Audit Trail"

DOC_COLUMNS = [
    "Doc. ID Number",
    "Document Description",
    "Fungsi",
    "Doc. Request Date",
    "Doc. Submission Deadline",
    "Days To Go",
    "Email - Auditee1",
    "Recepient - cc",
    "Document Status",
    "Remarks",
    "Status Reminder"
]

AUDIT_COLUMNS = ["Timestamp", "User", "Action", "Doc ID", "Details"]

USER_COLUMNS = ["Nama", "Email", "Role", "Fungsi"]

FUNGSI_COLUMNS = ["Nama Fungsi"]

DOC_STATUS_OPTIONS = ["Outstanding", "Need to Review", "Done"]

ROLE_OPTIONS = ["Admin", "User"]

DEFAULT_FUNGSI = [
    "Finance & Risk Management",
    "HC & Business Support",
    "Operation",
    "Legal & Compliance",
    "Information Technology",
    "Marketing & Sales",
    "Corporate Planning"
]

_DB_ENGINE = None
_DB_SCHEMA_READY = False


def _get_secret(key: str) -> Optional[str]:
    if st is not None:
        try:
            if key in st.secrets:
                value = st.secrets[key]
                return str(value) if value is not None else None
        except Exception:
            pass
    return os.getenv(key)


def get_database_url() -> Optional[str]:
    return _get_secret("SUPABASE_DB_URL")


def is_database_mode() -> bool:
    return bool(get_database_url())


def _get_db_engine():
    global _DB_ENGINE
    if _DB_ENGINE is not None:
        return _DB_ENGINE

    db_url = get_database_url()
    if not db_url:
        return None

    if create_engine is None:
        raise RuntimeError(
            "Database mode aktif, tapi dependency SQLAlchemy belum terpasang. "
            "Install: pip install sqlalchemy psycopg2-binary"
        )

    _DB_ENGINE = create_engine(db_url, pool_pre_ping=True)
    return _DB_ENGINE


def _coerce_datetime(value):
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    try:
        return pd.to_datetime(value).to_pydatetime()
    except Exception:
        return None


def _ensure_db_schema() -> None:
    global _DB_SCHEMA_READY
    if _DB_SCHEMA_READY:
        return

    engine = _get_db_engine()
    if engine is None:
        return

    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS documents (
                doc_id TEXT PRIMARY KEY,
                document_description TEXT,
                fungsi TEXT,
                doc_request_date TIMESTAMP NULL,
                doc_submission_deadline TIMESTAMP NULL,
                days_to_go INTEGER NULL,
                email_auditee1 TEXT,
                recipient_cc TEXT,
                document_status TEXT,
                remarks TEXT,
                status_reminder TEXT,
                created_at TIMESTAMPTZ DEFAULT NOW(),
                updated_at TIMESTAMPTZ DEFAULT NOW()
            );
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS users_data (
                id BIGSERIAL PRIMARY KEY,
                nama TEXT,
                email TEXT,
                role TEXT,
                fungsi TEXT
            );
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS master_fungsi (
                nama_fungsi TEXT PRIMARY KEY
            );
        """))

        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id BIGSERIAL PRIMARY KEY,
                timestamp TIMESTAMPTZ,
                username TEXT,
                action TEXT,
                doc_id TEXT,
                details TEXT
            );
        """))

        existing_count = conn.execute(text("SELECT COUNT(*) FROM master_fungsi;")).scalar() or 0
        if existing_count == 0:
            conn.execute(
                text("INSERT INTO master_fungsi (nama_fungsi) VALUES (:nama_fungsi);"),
                [{"nama_fungsi": fungsi_name} for fungsi_name in DEFAULT_FUNGSI]
            )

    _DB_SCHEMA_READY = True


def _db_read_documents() -> pd.DataFrame:
    _ensure_db_schema()
    engine = _get_db_engine()
    query = """
        SELECT
            doc_id AS "Doc. ID Number",
            document_description AS "Document Description",
            fungsi AS "Fungsi",
            doc_request_date AS "Doc. Request Date",
            doc_submission_deadline AS "Doc. Submission Deadline",
            days_to_go AS "Days To Go",
            email_auditee1 AS "Email - Auditee1",
            recipient_cc AS "Recepient - cc",
            document_status AS "Document Status",
            remarks AS "Remarks",
            status_reminder AS "Status Reminder"
        FROM documents
        ORDER BY doc_id;
    """
    df = pd.read_sql_query(query, engine)
    for col in DOC_COLUMNS:
        if col not in df.columns:
            df[col] = None
    if "Doc. Submission Deadline" in df.columns:
        df["Days To Go"] = df["Doc. Submission Deadline"].apply(calculate_days_to_go)
    return df[DOC_COLUMNS]


def _db_save_documents(df: pd.DataFrame) -> Tuple[bool, str]:
    try:
        _ensure_db_schema()
        engine = _get_db_engine()

        normalized = df.copy()
        for col in DOC_COLUMNS:
            if col not in normalized.columns:
                normalized[col] = None
        normalized = normalized[DOC_COLUMNS]

        if not normalized.empty:
            obj_cols = normalized.select_dtypes(include=['object']).columns
            for col in obj_cols:
                normalized[col] = normalized[col].apply(clean_input)

        rows = []
        for _, row in normalized.iterrows():
            doc_id = row.get("Doc. ID Number")
            if pd.isna(doc_id) or str(doc_id).strip() == "":
                continue
            rows.append({
                "doc_id": str(doc_id),
                "document_description": row.get("Document Description"),
                "fungsi": row.get("Fungsi"),
                "doc_request_date": _coerce_datetime(row.get("Doc. Request Date")),
                "doc_submission_deadline": _coerce_datetime(row.get("Doc. Submission Deadline")),
                "days_to_go": None if pd.isna(row.get("Days To Go")) else int(row.get("Days To Go")),
                "email_auditee1": row.get("Email - Auditee1"),
                "recipient_cc": row.get("Recepient - cc"),
                "document_status": row.get("Document Status"),
                "remarks": row.get("Remarks"),
                "status_reminder": row.get("Status Reminder")
            })

        with engine.begin() as conn:
            conn.execute(text("DELETE FROM documents;"))
            if rows:
                conn.execute(text("""
                    INSERT INTO documents (
                        doc_id, document_description, fungsi, doc_request_date,
                        doc_submission_deadline, days_to_go, email_auditee1,
                        recipient_cc, document_status, remarks, status_reminder,
                        created_at, updated_at
                    )
                    VALUES (
                        :doc_id, :document_description, :fungsi, :doc_request_date,
                        :doc_submission_deadline, :days_to_go, :email_auditee1,
                        :recipient_cc, :document_status, :remarks, :status_reminder,
                        NOW(), NOW()
                    );
                """), rows)

        return True, "✅ Data berhasil disimpan ke Supabase!"
    except Exception as e:
        return False, f"❌ Gagal simpan ke Supabase: {str(e)}"


def _db_read_users() -> pd.DataFrame:
    _ensure_db_schema()
    engine = _get_db_engine()
    query = """
        SELECT
            nama AS "Nama",
            email AS "Email",
            role AS "Role",
            fungsi AS "Fungsi"
        FROM users_data
        ORDER BY id;
    """
    df = pd.read_sql_query(query, engine)
    for col in USER_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[USER_COLUMNS]


def _db_save_users(df: pd.DataFrame) -> Tuple[bool, str]:
    try:
        _ensure_db_schema()
        engine = _get_db_engine()

        normalized = df.copy()
        for col in USER_COLUMNS:
            if col not in normalized.columns:
                normalized[col] = None
        normalized = normalized[USER_COLUMNS]

        if not normalized.empty:
            obj_cols = normalized.select_dtypes(include=['object']).columns
            for col in obj_cols:
                normalized[col] = normalized[col].apply(clean_input)

        rows = []
        for _, row in normalized.iterrows():
            rows.append({
                "nama": row.get("Nama"),
                "email": row.get("Email"),
                "role": row.get("Role"),
                "fungsi": row.get("Fungsi")
            })

        with engine.begin() as conn:
            conn.execute(text("DELETE FROM users_data;"))
            if rows:
                conn.execute(text("""
                    INSERT INTO users_data (nama, email, role, fungsi)
                    VALUES (:nama, :email, :role, :fungsi);
                """), rows)

        return True, "✅ Data user berhasil disimpan ke Supabase!"
    except Exception as e:
        return False, f"❌ Gagal simpan user ke Supabase: {str(e)}"


def _db_read_fungsi() -> pd.DataFrame:
    _ensure_db_schema()
    engine = _get_db_engine()
    query = """
        SELECT nama_fungsi AS "Nama Fungsi"
        FROM master_fungsi
        ORDER BY nama_fungsi;
    """
    df = pd.read_sql_query(query, engine)
    if df.empty:
        return pd.DataFrame({"Nama Fungsi": DEFAULT_FUNGSI})
    return df


def _db_save_fungsi(df: pd.DataFrame) -> Tuple[bool, str]:
    try:
        _ensure_db_schema()
        engine = _get_db_engine()

        normalized = df.copy()
        if "Nama Fungsi" not in normalized.columns:
            normalized["Nama Fungsi"] = None

        cleaned_values = []
        for value in normalized["Nama Fungsi"].tolist():
            if pd.isna(value):
                continue
            value_str = clean_input(str(value))
            if value_str:
                cleaned_values.append(value_str)

        unique_values = list(dict.fromkeys(cleaned_values))

        with engine.begin() as conn:
            conn.execute(text("DELETE FROM master_fungsi;"))
            if unique_values:
                conn.execute(
                    text("INSERT INTO master_fungsi (nama_fungsi) VALUES (:nama_fungsi);"),
                    [{"nama_fungsi": value} for value in unique_values]
                )

        return True, "✅ Data fungsi berhasil disimpan ke Supabase!"
    except Exception as e:
        return False, f"❌ Gagal simpan fungsi ke Supabase: {str(e)}"


def _db_read_audit_logs() -> pd.DataFrame:
    _ensure_db_schema()
    engine = _get_db_engine()
    query = """
        SELECT
            timestamp AS "Timestamp",
            username AS "User",
            action AS "Action",
            doc_id AS "Doc ID",
            details AS "Details"
        FROM audit_logs
        ORDER BY timestamp DESC;
    """
    df = pd.read_sql_query(query, engine)
    for col in AUDIT_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df[AUDIT_COLUMNS]


def _db_log_audit_event(user: str, action: str, doc_id: str, details: str = "") -> None:
    try:
        _ensure_db_schema()
        engine = _get_db_engine()
        with engine.begin() as conn:
            conn.execute(text("""
                INSERT INTO audit_logs (timestamp, username, action, doc_id, details)
                VALUES (:timestamp, :username, :action, :doc_id, :details);
            """), {
                "timestamp": datetime.now(),
                "username": user,
                "action": action,
                "doc_id": doc_id,
                "details": details
            })
    except Exception as e:
        print(f"Failed to log audit (db): {e}")

def get_base_path() -> str:
    """Returns the base path for resources."""
    if getattr(sys, 'frozen', False):
        # Running as compiled exe
        return sys._MEIPASS
    else:
        # Running as script
        return os.path.dirname(os.path.abspath(__file__))

def get_data_path() -> str:
    """Returns the path for persistent data (Excel file)."""
    if getattr(sys, 'frozen', False):
        # In exe mode, store data next to the executable
        return os.path.dirname(sys.executable)
    else:
        # In script mode, store data in current directory
        return os.path.dirname(os.path.abspath(__file__))

def get_excel_path() -> str:
    return os.path.join(get_data_path(), EXCEL_FILE)

def ensure_excel_exists() -> str:
    excel_path = get_excel_path()
    
    if not os.path.exists(excel_path):
        create_new_excel(excel_path)
    else:
        verify_and_repair_sheets(excel_path)
    
    return excel_path

def create_new_excel(excel_path: str) -> None:
    wb = Workbook()
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws_doc = wb.active
    ws_doc.title = SHEET_DOCUMENT
    for col, header in enumerate(DOC_COLUMNS, 1):
        cell = ws_doc.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws_doc.column_dimensions[chr(64 + col) if col < 27 else 'A' + chr(64 + col - 26)].width = 20
    
    ws_users = wb.create_sheet(SHEET_USERS)
    for col, header in enumerate(USER_COLUMNS, 1):
        cell = ws_users.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws_users.column_dimensions[chr(64 + col)].width = 25
    
    ws_fungsi = wb.create_sheet(SHEET_FUNGSI)
    for col, header in enumerate(FUNGSI_COLUMNS, 1):
        cell = ws_fungsi.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws_fungsi.column_dimensions['A'].width = 35
    
    for row, fungsi in enumerate(DEFAULT_FUNGSI, 2):
        ws_fungsi.cell(row=row, column=1, value=fungsi)
    
    ws_audit = wb.create_sheet(SHEET_AUDIT)
    for col, header in enumerate(AUDIT_COLUMNS, 1):
        cell = ws_audit.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws_audit.column_dimensions['A'].width = 20
    ws_audit.column_dimensions['B'].width = 20
    ws_audit.column_dimensions['C'].width = 15
    ws_audit.column_dimensions['D'].width = 15
    ws_audit.column_dimensions['E'].width = 50
    
    wb.save(excel_path)
    wb.close()

def verify_and_repair_sheets(excel_path: str) -> None:
    try:
        wb = load_workbook(excel_path)
        sheet_names = wb.sheetnames
        
        needs_save = False
        
        if SHEET_DOCUMENT not in sheet_names:
            ws = wb.create_sheet(SHEET_DOCUMENT)
            for col, header in enumerate(DOC_COLUMNS, 1):
                ws.cell(row=1, column=col, value=header)
            needs_save = True
        
        if SHEET_USERS not in sheet_names:
            ws = wb.create_sheet(SHEET_USERS)
            for col, header in enumerate(USER_COLUMNS, 1):
                ws.cell(row=1, column=col, value=header)
            needs_save = True
        
        if SHEET_FUNGSI not in sheet_names:
            ws = wb.create_sheet(SHEET_FUNGSI)
            ws.cell(row=1, column=1, value="Nama Fungsi")
            for row, fungsi in enumerate(DEFAULT_FUNGSI, 2):
                ws.cell(row=row, column=1, value=fungsi)
            needs_save = True

        if SHEET_AUDIT not in sheet_names:
            ws = wb.create_sheet(SHEET_AUDIT)
            for col, header in enumerate(AUDIT_COLUMNS, 1):
                ws.cell(row=1, column=col, value=header)
            needs_save = True
        
        if needs_save:
            wb.save(excel_path)
        wb.close()
        
    except PermissionError:
        raise PermissionError(
            "❌ File Excel sedang dibuka di aplikasi lain. "
            "Tutup file tersebut terlebih dahulu."
        )

def load_documents() -> pd.DataFrame:
    if is_database_mode():
        try:
            return _db_read_documents()
        except Exception:
            return pd.DataFrame(columns=DOC_COLUMNS)

    try:
        excel_path = ensure_excel_exists()
        df = pd.read_excel(excel_path, sheet_name=SHEET_DOCUMENT, engine='openpyxl')
        
        for col in DOC_COLUMNS:
            if col not in df.columns:
                df[col] = None
        
        if 'Doc. Submission Deadline' in df.columns:
            df['Days To Go'] = df['Doc. Submission Deadline'].apply(calculate_days_to_go)
        
        return df
        
    except Exception as e:
        return pd.DataFrame(columns=DOC_COLUMNS)

def process_bulk_upload(uploaded_file) -> Tuple[bool, str, int]:
    try:
        df_upload = pd.read_excel(uploaded_file, engine='openpyxl')
        
        # Mapping kolom fleksibel (case insensitive)
        req_cols = {
            'document description': 'Document Description',
            'fungsi': 'Fungsi',
            'divisi': 'Fungsi',
            'request date': 'Doc. Request Date',
            'deadline': 'Doc. Submission Deadline',
            'email': 'Email - Auditee1',
            'status': 'Document Status'
        }
        
        df_upload.columns = df_upload.columns.str.strip()
        
        # Rename kolom yang dikenali
        cols_map = {}
        for col in df_upload.columns:
            if col.lower() in req_cols:
                cols_map[col] = req_cols[col.lower()]
        
        df_upload.rename(columns=cols_map, inplace=True)
        
        # Validasi kolom wajib
        missing_cols = [c for c in ['Document Description', 'Fungsi', 'Doc. Submission Deadline'] if c not in df_upload.columns]
        if missing_cols:
            return False, f"❌ Kolom wajib tidak ditemukan: {', '.join(missing_cols)}", 0
            
        success_count = 0
        df_existing = load_documents()
        
        new_rows = []
        for index, row in df_upload.iterrows():
            new_id = generate_doc_id("BULK")
            # Cek duplikasi ID di batch ini
            while any(r['Doc. ID Number'] == new_id for r in new_rows) or new_id in df_existing['Doc. ID Number'].values:
                # Increment manual jika konflik
                num_part = int(new_id.split('-')[-1]) + 1
                prefix_year = "-".join(new_id.split('-')[:2])
                new_id = f"{prefix_year}-{num_part:03d}"
            
            row_data = {
                "Doc. ID Number": row.get('Doc. ID Number', new_id), # Gunakan ID dari excel jika ada
                "Document Description": row['Document Description'],
                "Fungsi": row['Fungsi'],
                "Doc. Request Date": row.get('Doc. Request Date', datetime.now()),
                "Doc. Submission Deadline": row['Doc. Submission Deadline'],
                "Days To Go": calculate_days_to_go(row['Doc. Submission Deadline']), # Calculated
                "Email - Auditee1": row.get('Email - Auditee1', None),
                "Recepient - cc": row.get('Recepient - cc', None),
                "Document Status": row.get('Document Status', 'Outstanding'),
                "Remarks": row.get('Remarks', 'Bulk Upload'),
                "Status Reminder": None
            }
            new_rows.append(row_data)
            success_count += 1

        if new_rows:
            df_new = pd.DataFrame(new_rows)
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            saved, msg = save_documents(df_combined)
            if saved:
                 log_audit_event("System", "Bulk Upload", "Batch", f"Uploaded {success_count} documents")
                 return True, f"✅ Berhasil import {success_count} dokumen!", success_count
            else:
                 return False, msg, 0
        
        return False, "⚠️ Tidak ada data yang bisa diproses.", 0
        
    except Exception as e:
        return False, f"❌ Error membaca file: {str(e)}", 0

def load_users() -> pd.DataFrame:
    if is_database_mode():
        try:
            return _db_read_users()
        except Exception:
            return pd.DataFrame(columns=USER_COLUMNS)

    excel_path = ensure_excel_exists()
    
    try:
        df = pd.read_excel(excel_path, sheet_name=SHEET_USERS, engine='openpyxl')
        for col in USER_COLUMNS:
            if col not in df.columns:
                df[col] = None
        return df
    except Exception:
        return pd.DataFrame(columns=USER_COLUMNS)

def load_fungsi() -> pd.DataFrame:
    if is_database_mode():
        try:
            return _db_read_fungsi()
        except Exception:
            return pd.DataFrame(columns=FUNGSI_COLUMNS)

    excel_path = ensure_excel_exists()
    
    try:
        df = pd.read_excel(excel_path, sheet_name=SHEET_FUNGSI, engine='openpyxl')
        return df
    except Exception:
        return pd.DataFrame(columns=FUNGSI_COLUMNS)

def get_fungsi_list() -> List[str]:
    df = load_fungsi()
    if 'Nama Fungsi' in df.columns:
        return df['Nama Fungsi'].dropna().tolist()
    return DEFAULT_FUNGSI

def load_audit_logs() -> pd.DataFrame:
    if is_database_mode():
        try:
            return _db_read_audit_logs()
        except Exception:
            return pd.DataFrame(columns=AUDIT_COLUMNS)

    excel_path = ensure_excel_exists()
    try:
        df = pd.read_excel(excel_path, sheet_name=SHEET_AUDIT, engine='openpyxl')
        if not df.empty and 'Timestamp' in df.columns:
            df = df.sort_values(by='Timestamp', ascending=False)
        return df
    except Exception:
         return pd.DataFrame(columns=AUDIT_COLUMNS)

def log_audit_event(user: str, action: str, doc_id: str, details: str = "") -> None:
    if is_database_mode():
        _db_log_audit_event(user, action, doc_id, details)
        return

    try:
        excel_path = ensure_excel_exists()
        wb = load_workbook(excel_path)
        
        if SHEET_AUDIT not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_AUDIT)
            ws.append(AUDIT_COLUMNS)
        else:
            ws = wb[SHEET_AUDIT]
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([timestamp, user, action, doc_id, details])
        
        wb.save(excel_path)
        wb.close()
    except Exception as e:
        print(f"Failed to log audit: {e}")

def clean_input(text: str) -> str:
    if not isinstance(text, str):
        return text
    # Hapus karakter kontrol Excel yang berbahaya
    return re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', text).strip()

def save_to_excel(df: pd.DataFrame, sheet_name: str, retry_count: int = 5) -> Tuple[bool, str]:
    excel_path = ensure_excel_exists()
    
    # Bersihkan input sebelum simpan
    if not df.empty:
        obj_cols = df.select_dtypes(include=['object']).columns
        for col in obj_cols:
            df[col] = df[col].apply(clean_input)

    for attempt in range(retry_count):
        try:
            wb = load_workbook(excel_path)
            
            if sheet_name in wb.sheetnames:
                # Hapus sheet lama
                del wb[sheet_name]
            
            ws = wb.create_sheet(sheet_name)
            
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for col, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    if pd.isna(value):
                        cell_value = None
                    elif isinstance(value, pd.Timestamp):
                        cell_value = value.to_pydatetime()
                    else:
                        cell_value = value
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Sembunyikan sheet Audit Trail dan Users agar tidak mudah diubah manual
            if sheet_name in [SHEET_AUDIT, SHEET_USERS]:
                ws.sheet_state = 'hidden'

            wb.save(excel_path)
            wb.close()
            return True, "✅ Data berhasil disimpan!"
            
        except PermissionError:
            if attempt < retry_count - 1:
                time.sleep(2) # Tunggu 2 detik sebelum coba lagi (SharePoint sync delay)
                continue
            return False, "❌ File Excel sedang disinkronisasi atau dibuka user lain. Coba sesaat lagi."
        except Exception as e:
            return False, f"❌ Error: {str(e)}"
    
    return False, "❌ Gagal menyimpan setelah beberapa percobaan."

def save_documents(df: pd.DataFrame) -> Tuple[bool, str]:
    if is_database_mode():
        return _db_save_documents(df)
    return save_to_excel(df, SHEET_DOCUMENT)

def save_users(df: pd.DataFrame) -> Tuple[bool, str]:
    if is_database_mode():
        return _db_save_users(df)
    return save_to_excel(df, SHEET_USERS)

def save_fungsi(df: pd.DataFrame) -> Tuple[bool, str]:
    if is_database_mode():
        return _db_save_fungsi(df)
    return save_to_excel(df, SHEET_FUNGSI)

def add_document(doc_data: dict) -> Tuple[bool, str]:
    df = load_documents()
    
    if doc_data.get('Doc. ID Number') in df['Doc. ID Number'].values:
        return False, "❌ Doc. ID Number sudah ada! Gunakan ID yang berbeda."
    
    new_row = pd.DataFrame([doc_data])
    df = pd.concat([df, new_row], ignore_index=True)
    
    success, message = save_documents(df)
    if success:
        log_audit_event("User", "Create", doc_data.get('Doc. ID Number'), "New document created")
    return success, message

def update_document(doc_id: str, updates: dict) -> Tuple[bool, str]:
    df = load_documents()
    
    if doc_id not in df['Doc. ID Number'].values:
        return False, "❌ Dokumen tidak ditemukan!"
    
    idx = df[df['Doc. ID Number'] == doc_id].index[0]
    
    changes = []
    for key, value in updates.items():
        if key in df.columns:
            old_val = df.at[idx, key]
            if str(old_val) != str(value):
                changes.append(f"{key}: {old_val} -> {value}")
            df.at[idx, key] = value
    
    success, message = save_documents(df)
    if success and changes:
        log_audit_event("User", "Update", doc_id, ", ".join(changes))
    return success, message

def delete_document(doc_id: str) -> Tuple[bool, str]:
    df = load_documents()
    
    if doc_id not in df['Doc. ID Number'].values:
        return False, "❌ Dokumen tidak ditemukan!"
    
    df = df[df['Doc. ID Number'] != doc_id]
    
    success, message = save_documents(df)
    if success:
        log_audit_event("User", "Delete", doc_id, "Document deleted")
    return success, message

def update_reminder_status(doc_id: str) -> Tuple[bool, str]:
    today = datetime.now().strftime("%Y-%m-%d")
    return update_document(doc_id, {"Status Reminder": f"Sent: {today}"})


def calculate_days_to_go(deadline) -> Optional[int]:
    if pd.isna(deadline):
        return None
    
    try:
        if isinstance(deadline, str):
            deadline = pd.to_datetime(deadline)
        elif isinstance(deadline, datetime):
            pass
        elif isinstance(deadline, date):
            deadline = datetime.combine(deadline, datetime.min.time())
        else:
            return None
        
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        if isinstance(deadline, pd.Timestamp):
            deadline = deadline.to_pydatetime()
        deadline = deadline.replace(hour=0, minute=0, second=0, microsecond=0)
        
        delta = (deadline - today).days
        return delta
    except:
        return None

def process_bulk_upload(uploaded_file) -> Tuple[bool, str, int]:
    try:
        df_upload = pd.read_excel(uploaded_file, engine='openpyxl')
        
        # Mapping kolom fleksibel (case insensitive)
        req_cols = {
            'document description': 'Document Description',
            'fungsi': 'Fungsi',
            'divisi': 'Fungsi',
            'request date': 'Doc. Request Date',
            'deadline': 'Doc. Submission Deadline',
            'email': 'Email - Auditee1',
            'status': 'Document Status'
        }
        
        df_upload.columns = df_upload.columns.str.strip()
        
        # Rename kolom yang dikenali
        cols_map = {}
        for col in df_upload.columns:
            if col.lower() in req_cols:
                cols_map[col] = req_cols[col.lower()]
        
        df_upload.rename(columns=cols_map, inplace=True)
        
        # Validasi kolom wajib
        missing_cols = [c for c in ['Document Description', 'Fungsi', 'Doc. Submission Deadline'] if c not in df_upload.columns]
        if missing_cols:
            return False, f"❌ Kolom wajib tidak ditemukan: {', '.join(missing_cols)}", 0
            
        success_count = 0
        df_existing = load_documents()
        
        new_rows = []
        for index, row in df_upload.iterrows():
            new_id = generate_doc_id("BULK")
            # Cek duplikasi ID di batch ini
            while any(r['Doc. ID Number'] == new_id for r in new_rows) or new_id in df_existing['Doc. ID Number'].values:
                # Increment manual jika konflik
                num_part = int(new_id.split('-')[-1]) + 1
                prefix_year = "-".join(new_id.split('-')[:2])
                new_id = f"{prefix_year}-{num_part:03d}"
            
            row_data = {
                "Doc. ID Number": row.get('Doc. ID Number', new_id), # Gunakan ID dari excel jika ada
                "Document Description": row['Document Description'],
                "Fungsi": row['Fungsi'],
                "Doc. Request Date": row.get('Doc. Request Date', datetime.now()),
                "Doc. Submission Deadline": row['Doc. Submission Deadline'],
                "Days To Go": None, # Will be calc automatically
                "Email - Auditee1": row.get('Email - Auditee1', None),
                "Recepient - cc": row.get('Recepient - cc', None),
                "Document Status": row.get('Document Status', 'Outstanding'),
                "Remarks": row.get('Remarks', 'Bulk Upload'),
                "Status Reminder": None
            }
            new_rows.append(row_data)
            success_count += 1

        if new_rows:
            df_new = pd.DataFrame(new_rows)
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            saved, msg = save_documents(df_combined)
            if saved:
                 log_audit_event("System", "Bulk Upload", "Batch", f"Uploaded {success_count} documents")
                 return True, f"✅ Berhasil import {success_count} dokumen!", success_count
            else:
                 return False, msg, 0
        
        return False, "⚠️ Tidak ada data yang bisa diproses.", 0
        
    except Exception as e:
        return False, f"❌ Error membaca file: {str(e)}", 0

def get_status_color(status: str) -> str:
    status_colors = {
        "Outstanding": "#DA291C",
        "Need to Review": "#FFC107",
        "Done": "#469B43"
    }
    return status_colors.get(status, "#003366")

def get_days_color(days: int) -> str:
    if days is None:
        return "#6c757d"
    elif days < 0:
        return "#DA291C"
    elif days <= 3:
        return "#FFC107"
    else:
        return "#469B43"

def generate_doc_id(prefix: str = "DOC") -> str:
    df = load_documents()
    year = datetime.now().year
    
    pattern = f"{prefix}-{year}-"
    existing = df[df['Doc. ID Number'].str.startswith(pattern, na=False)]
    
    if existing.empty:
        next_num = 1
    else:
        numbers = existing['Doc. ID Number'].str.extract(r'(\d+)$')[0].astype(int)
        next_num = numbers.max() + 1
    
    return f"{prefix}-{year}-{next_num:03d}"

def get_documents_for_reminder() -> pd.DataFrame:
    df = load_documents()
    
    reminder_status = ['Outstanding', 'Need to Review']
    filtered = df[df['Document Status'].isin(reminder_status)]
    
    return filtered

def generate_email_body(doc_data: dict) -> str:
    doc_id = doc_data.get('Doc. ID Number', '-')
    doc_desc = doc_data.get('Document Description', '-')
    deadline = doc_data.get('Doc. Submission Deadline', '')
    days_to_go = doc_data.get('Days To Go', 0)
    status = doc_data.get('Document Status', '-')
    
    if isinstance(deadline, (datetime, date)):
        deadline_str = deadline.strftime("%d %B %Y")
    elif isinstance(deadline, pd.Timestamp):
        deadline_str = deadline.strftime("%d %B %Y")
    else:
        deadline_str = str(deadline) if deadline else "-"
    
    if days_to_go is not None and days_to_go < 0:
        urgency = f"OVERDUE ({abs(days_to_go)} hari)"
    elif days_to_go is not None and days_to_go <= 3:
        urgency = f"Mendesak ({days_to_go} hari lagi)"
    else:
        urgency = f"{days_to_go} hari lagi" if days_to_go else "-"

    body = f"""Yth. Rekan Auditee,

Semoga Bapak/Ibu dalam keadaan baik.

Melalui email ini, kami mengingatkan tindak lanjut dokumen audit dengan detail sebagai berikut:

- No. Dokumen       : {doc_id}
- Deskripsi         : {doc_desc}
- Status            : {status}
- Deadline          : {deadline_str}
- Sisa Waktu        : {urgency}

Mohon dukungan Bapak/Ibu untuk melengkapi dokumen dan menyampaikan respon melalui link Microsoft Form yang berlaku.

Apabila dokumen telah dikirim, mohon abaikan pengingat ini.

Terima kasih atas perhatian dan kerja samanya.

Hormat kami,
Tim Audit Internal
Dashboard Audit - Pertamina Corporate
"""
    return body

def generate_email_subject(doc_data: dict) -> str:
    status = doc_data.get('Document Status', 'Outstanding')
    doc_desc = doc_data.get('Document Description', 'Dokumen Audit')
    return f"REMINDER: [{status}] - {doc_desc}"

def generate_outlook_email(
    to: str,
    cc: str,
    subject: str,
    body_html: str,
    doc_id: str = None
) -> Tuple[bool, str]:
    pythoncom_initialized = False
    try:
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        pythoncom_initialized = True
        
        outlook = win32com.client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        
        mail.To = to
        mail.CC = cc if cc else ""
        mail.Subject = subject
        mail.HTMLBody = body_html
        
        mail.Display(True)
        
        if doc_id:
            update_reminder_status(doc_id)
        
        return True, "✅ Email berhasil dibuat di Outlook!"
        
    except ImportError:
        return False, (
            "❌ Module pywin32 tidak tersedia. Fitur Outlook COM hanya berjalan di Windows desktop. "
            "Untuk Streamlit Cloud gunakan layanan email server-side (Microsoft Graph/SMTP)."
        )
    except Exception as e:
        error_msg = str(e)
        error_msg_lower = error_msg.lower()
        if "coinitialize has not been called" in error_msg_lower or "-2147221008" in error_msg:
            return False, (
                "❌ Outlook COM belum siap (`CoInitialize has not been called`). "
                "Tutup aplikasi, buka Outlook Desktop secara manual, lalu coba lagi."
            )
        if "Outlook" in error_msg or "dispatch" in error_msg.lower():
            return False, "❌ Outlook Desktop tidak ditemukan atau tidak aktif."
        return False, f"❌ Error: {error_msg}"
    finally:
        if pythoncom_initialized:
            try:
                import pythoncom
                pythoncom.CoUninitialize()
            except Exception:
                pass

def _html_to_plain_text(html_body: str) -> str:
    if not html_body:
        return ""

    text = re.sub(r"(?i)<br\s*/?>", "\n", html_body)
    text = re.sub(r"(?i)</p>", "\n\n", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()

def generate_outlook_email_new(
    to: str,
    cc: str,
    subject: str,
    body_html: str,
    doc_id: str = None
) -> Tuple[bool, str]:
    try:
        body_text = _html_to_plain_text(body_html)
        if len(body_text) > 1800:
            body_text = body_text[:1800] + "\n\n..."

        mailto_url = (
            f"mailto:{quote(to or '')}"
            f"?cc={quote(cc or '')}&subject={quote(subject or '')}&body={quote(body_text or '')}"
        )

        opened = False
        try:
            os.startfile(mailto_url)
            opened = True
        except Exception:
            pass

        if not opened:
            import webbrowser
            opened = webbrowser.open(mailto_url)

        if not opened:
            return False, "❌ Tidak bisa membuka compose email. Pastikan New Outlook sudah dijadikan default mail app di Windows."

        if doc_id:
            update_reminder_status(doc_id)

        return True, "✅ Jendela compose email berhasil dibuka di aplikasi Outlook default."
    except Exception as e:
        return False, f"❌ Gagal membuka Outlook terbaru: {str(e)}"

def get_document_statistics() -> dict:
    df = load_documents()
    
    total = len(df)
    outstanding = len(df[df['Document Status'] == 'Outstanding'])
    need_review = len(df[df['Document Status'] == 'Need to Review'])
    done = len(df[df['Document Status'] == 'Done'])
    
    overdue = len(df[df['Days To Go'] < 0]) if 'Days To Go' in df.columns else 0
    
    return {
        'total': total,
        'outstanding': outstanding,
        'need_review': need_review,
        'done': done,
        'overdue': overdue,
        'completion_rate': round((done / total * 100) if total > 0 else 0, 1)
    }

def get_status_by_fungsi() -> pd.DataFrame:
    df = load_documents()
    
    if df.empty:
        return pd.DataFrame()
    


    stats = df.groupby(['Fungsi', 'Document Status']).size().unstack(fill_value=0)
    stats = stats.reset_index()
    
    return stats

def generate_spectacular_report(df_docs: pd.DataFrame = None) -> bytes:
    output = BytesIO()
    wb = Workbook()
    
    if df_docs is None:
        df_docs = load_documents()

    # Calculate dynamic stats based on filtered data
    total = len(df_docs)
    outstanding = len(df_docs[df_docs['Document Status'] == 'Outstanding'])
    need_review = len(df_docs[df_docs['Document Status'] == 'Need to Review'])
    done = len(df_docs[df_docs['Document Status'] == 'Done'])
    
    # Calculate overdue based on filtered data
    if 'Days To Go' in df_docs.columns:
        overdue = len(df_docs[df_docs['Days To Go'] < 0])
    else:
        # Fallback calculation if column missing
        try:
           today = datetime.now().date()
           overdue = len(df_docs[df_docs['Doc. Submission Deadline'].apply(lambda x: pd.to_datetime(x).date() < today if pd.notnull(x) else False)])
        except:
           overdue = 0

    completion_rate = round((done / total * 100) if total > 0 else 0, 1)

    # --- STYLES ---
    title_font = Font(name='Calibri', size=20, bold=True, color="003366")
    subtitle_font = Font(name='Calibri', size=12, italic=True, color="555555")
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    text_font = Font(name='Calibri', size=11)
    bold_font = Font(name='Calibri', size=11, bold=True)
    
    fill_header = PatternFill(start_color="005DAA", end_color="005DAA", fill_type="solid")
    fill_light = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
    
    border_thin = Border(left=Side(style='thin', color="CCCCCC"), 
                        right=Side(style='thin', color="CCCCCC"), 
                        top=Side(style='thin', color="CCCCCC"), 
                        bottom=Side(style='thin', color="CCCCCC"))
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- SHEET 1: EXECUTIVE SUMMARY ---
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    ws_summary.sheet_view.showGridLines = False
    
    # Title Section
    ws_summary.merge_cells('B2:E2')
    ws_summary['B2'] = "AUDIT DASHBOARD REPORT"
    ws_summary['B2'].font = title_font
    ws_summary['B2'].alignment = Alignment(horizontal="left", vertical="center")
    
    ws_summary.merge_cells('B3:E3')
    ws_summary['B3'] = f"Generated on: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    ws_summary['B3'].font = subtitle_font
    ws_summary['B3'].alignment = Alignment(horizontal="left", vertical="center")

    # Metrics Grid
    metrics = [
        ("Total Documents", total, "Total dokumen audit dalam periode ini"),
        ("Completion Rate", f"{completion_rate}%", "Persentase penyelesaian dokumen"),
        ("Outstanding", outstanding, "Dokumen belum diproses"),
        ("On Progress", need_review, "Dokumen sedang direview"),
        ("Completed", done, "Dokumen selesai (Done)"),
        ("Overdue/Late", overdue, "Dokumen melewati batas waktu")
    ]

    start_row = 6
    ws_summary['B5'] = "Key Performance Indicators"
    ws_summary['B5'].font = Font(name='Calibri', size=14, bold=True, color="005DAA")

    # Header Row for Metrics
    headers = ["Metric Indicator", "Value", "Description"]
    for col_idx, header in enumerate(headers, 2):
        cell = ws_summary.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # Data Rows for Metrics
    for i, (name, value, desc) in enumerate(metrics):
        row_num = start_row + 1 + i
        
        # Name
        cell_name = ws_summary.cell(row=row_num, column=2, value=name)
        cell_name.font = bold_font
        cell_name.border = border_thin
        cell_name.fill = fill_light
        
        # Value
        cell_val = ws_summary.cell(row=row_num, column=3, value=value)
        cell_val.font = text_font
        cell_val.border = border_thin
        cell_val.alignment = align_center
        
        # Conditional Formatting for Values
        if name == "Outstanding" and value > 0:
            cell_val.font = Font(color="DA291C", bold=True)
        elif name == "Completed":
            cell_val.font = Font(color="469B43", bold=True)
        elif name == "Overdue/Late" and value > 0:
            cell_val.font = Font(color="DA291C", bold=True)
            
        # Description
        cell_desc = ws_summary.cell(row=row_num, column=4, value=desc)
        cell_desc.font = Font(name='Calibri', size=11, italic=True, color="555555")
        cell_desc.border = border_thin

    # Division Breakdown Section
    breakdown_start = start_row + len(metrics) + 4
    ws_summary.cell(row=breakdown_start-1, column=2, value="Breakdown by Division").font = Font(name='Calibri', size=14, bold=True, color="005DAA")
    
    # Division Headers
    div_headers = ["Division / Fungsi", "Total Docs", "Completed", "Outstanding", "% Progress"]
    for col_idx, header in enumerate(div_headers, 2):
        cell = ws_summary.cell(row=breakdown_start, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # Division Data
    if not df_docs.empty:
        div_stats = df_docs.groupby('Fungsi').agg({
            'Doc. ID Number': 'count',
            'Document Status': lambda x: (x == 'Done').sum()
        }).reset_index()
        div_stats.columns = ['Fungsi', 'Total', 'Done']
        div_stats['Outstanding'] = div_stats['Total'] - div_stats['Done']
        div_stats['Progress'] = (div_stats['Done'] / div_stats['Total'] * 100).round(1)
        
        for i, row in enumerate(div_stats.itertuples(index=False)):
            r = breakdown_start + 1 + i
            ws_summary.cell(row=r, column=2, value=row.Fungsi).border = border_thin
            ws_summary.cell(row=r, column=3, value=row.Total).border = border_thin
            ws_summary.cell(row=r, column=3).alignment = align_center
            
            ws_summary.cell(row=r, column=4, value=row.Done).border = border_thin
            ws_summary.cell(row=r, column=4).alignment = align_center
            ws_summary.cell(row=r, column=4).font = Font(color="469B43", bold=True)
            
            ws_summary.cell(row=r, column=5, value=row.Outstanding).border = border_thin
            ws_summary.cell(row=r, column=5).alignment = align_center
            if row.Outstanding > 0:
                 ws_summary.cell(row=r, column=5).font = Font(color="DA291C")
            
            ws_summary.cell(row=r, column=6, value=f"{row.Progress}%").border = border_thin
            ws_summary.cell(row=r, column=6).alignment = align_center

    # Column Widths
    ws_summary.column_dimensions['A'].width = 2
    ws_summary.column_dimensions['B'].width = 35
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 20
    ws_summary.column_dimensions['E'].width = 15

    # --- SHEET 2: DETAILED DATA ---
    ws_data = wb.create_sheet("Detailed Data")
    
    # Headers
    headers = list(df_docs.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # Data Rows
    for r_idx, row in enumerate(df_docs.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
            cell.font = text_font
            cell.border = border_thin
            cell.alignment = align_left
            
            # Format Dates
            if isinstance(value, (datetime, pd.Timestamp)):
                cell.number_format = 'DD MMM YYYY'
                
            # Conditional Formatting for Status
            if headers[c_idx-1] == "Document Status":
                cell.alignment = align_center
                cell.font = bold_font
                if value == "Outstanding":
                    cell.font = Font(color="DA291C", bold=True)
                elif value == "Done":
                    cell.font = Font(color="469B43", bold=True)
                elif value == "Need to Review":
                    cell.font = Font(color="FFC107", bold=True)

    # Auto-fit Columns
    for col in ws_data.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_data.column_dimensions[column].width = min(adjusted_width, 50)

    wb.save(output)
    return output.getvalue()
